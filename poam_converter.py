"""
poam_converter.py — CLI tool that converts vulnerability scanner exports
into a FedRAMP-compliant POA&M Excel file.

Supported scanners: Nessus, Tenable, Qualys, Wiz, Generic CSV

Usage:
    python3 poam_converter.py --input scan.csv --scanner nessus --output poam.xlsx
    python3 poam_converter.py --input scan.csv --scanner qualys --output poam.xlsx
    python3 poam_converter.py --input scan.csv --scanner wiz --output poam.xlsx
    python3 poam_converter.py --input scan.csv --scanner generic --output poam.xlsx
    python3 poam_converter.py --input scan.xml --scanner nessus --output poam.xlsx
"""

import argparse
import csv
import sys
import xml.etree.ElementTree as ET
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Normalized finding schema
# ---------------------------------------------------------------------------

def empty_finding():
    return {
        "finding_id": "",
        "title": "",
        "severity": "",
        "cve": "",
        "asset": "",
        "description": "",
        "solution": "",
        "first_seen": "",
        "plugin_id": "",
        "controls": "",
    }


# ---------------------------------------------------------------------------
# Scanner parsers — each returns a list of normalized findings
# ---------------------------------------------------------------------------

def parse_nessus_csv(filepath):
    findings = []
    with open(filepath, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for i, row in enumerate(reader):
            sev = row.get("Risk", row.get("Severity", "")).strip()
            if not sev or sev.lower() in ("none", "informational", "info"):
                continue
            f_ = empty_finding()
            f_["finding_id"] = f"NESSUS-{i+1:04d}"
            f_["title"] = row.get("Name", row.get("Plugin Name", ""))
            f_["severity"] = normalize_severity(sev)
            f_["cve"] = row.get("CVE", "")
            f_["asset"] = row.get("Host", row.get("IP Address", ""))
            f_["description"] = row.get("Synopsis", row.get("Description", ""))
            f_["solution"] = row.get("Solution", "")
            f_["first_seen"] = row.get("First Discovered", row.get("Discovery Date", ""))
            f_["plugin_id"] = row.get("Plugin ID", "")
            findings.append(f_)
    return findings


def parse_nessus_xml(filepath):
    findings = []
    tree = ET.parse(filepath)
    root = tree.getroot()
    counter = 1
    for report_host in root.iter("ReportHost"):
        host = report_host.get("name", "")
        for item in report_host.iter("ReportItem"):
            sev_num = item.get("severity", "0")
            if sev_num in ("0",):
                continue
            sev_map = {"1": "Low", "2": "Medium", "3": "High", "4": "Critical"}
            sev = sev_map.get(sev_num, "Informational")
            f_ = empty_finding()
            f_["finding_id"] = f"NESSUS-{counter:04d}"
            f_["title"] = item.get("pluginName", "")
            f_["severity"] = sev
            cve_el = item.find("cve")
            f_["cve"] = cve_el.text if cve_el is not None else ""
            f_["asset"] = host
            syn = item.find("synopsis")
            f_["description"] = syn.text if syn is not None else ""
            sol = item.find("solution")
            f_["solution"] = sol.text if sol is not None else ""
            f_["plugin_id"] = item.get("pluginID", "")
            findings.append(f_)
            counter += 1
    return findings


def parse_tenable_csv(filepath):
    findings = []
    with open(filepath, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for i, row in enumerate(reader):
            sev = row.get("Severity", row.get("Risk Factor", "")).strip()
            if not sev or sev.lower() in ("none", "info", "informational"):
                continue
            f_ = empty_finding()
            f_["finding_id"] = f"TENABLE-{i+1:04d}"
            f_["title"] = row.get("Name", row.get("Plugin Name", ""))
            f_["severity"] = normalize_severity(sev)
            f_["cve"] = row.get("CVE", "")
            f_["asset"] = row.get("Asset", row.get("IP Address", row.get("Host", "")))
            f_["description"] = row.get("Description", row.get("Synopsis", ""))
            f_["solution"] = row.get("Solution", "")
            f_["first_seen"] = row.get("First Seen", row.get("First Discovered", ""))
            f_["plugin_id"] = row.get("Plugin ID", "")
            findings.append(f_)
    return findings


def parse_qualys_csv(filepath):
    findings = []
    with open(filepath, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for i, row in enumerate(reader):
            sev = row.get("Severity", row.get("Severity Level", "")).strip()
            if not sev or sev in ("1", "2"):  # Qualys 1-2 = Info/Low-ish
                continue
            f_ = empty_finding()
            f_["finding_id"] = row.get("QID", f"QUALYS-{i+1:04d}")
            f_["title"] = row.get("Title", row.get("Vulnerability", ""))
            f_["severity"] = normalize_severity_qualys(sev)
            f_["cve"] = row.get("CVE ID", row.get("CVE", ""))
            f_["asset"] = row.get("IP", row.get("Asset IP", row.get("Host", "")))
            f_["description"] = row.get("Threat", row.get("Description", ""))
            f_["solution"] = row.get("Solution", row.get("Remediation", ""))
            f_["first_seen"] = row.get("First Detected", row.get("First Found", ""))
            findings.append(f_)
    return findings


def parse_wiz_csv(filepath):
    findings = []
    with open(filepath, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for i, row in enumerate(reader):
            sev = row.get("Severity", row.get("Risk", "")).strip()
            if not sev or sev.lower() in ("info", "informational", "low"):
                continue
            f_ = empty_finding()
            f_["finding_id"] = row.get("ID", row.get("Issue ID", f"WIZ-{i+1:04d}"))
            f_["title"] = row.get("Title", row.get("Issue Title", row.get("Name", "")))
            f_["severity"] = normalize_severity(sev)
            f_["cve"] = row.get("CVE", row.get("CVE IDs", ""))
            f_["asset"] = row.get("Resource Name", row.get("Asset", row.get("Resource", "")))
            f_["description"] = row.get("Description", row.get("Details", ""))
            f_["solution"] = row.get("Remediation", row.get("Resolution", row.get("Fix", "")))
            f_["first_seen"] = row.get("Created At", row.get("First Detected", ""))
            findings.append(f_)
    return findings


def parse_generic_csv(filepath):
    """Fallback parser — tries common field name variations."""
    findings = []
    with open(filepath, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames or []
        for i, row in enumerate(reader):
            sev = (
                row.get("Severity") or row.get("Risk") or
                row.get("Risk Level") or row.get("Priority") or ""
            ).strip()
            if not sev or sev.lower() in ("none", "info", "informational"):
                continue
            f_ = empty_finding()
            f_["finding_id"] = (
                row.get("ID") or row.get("Finding ID") or
                row.get("Vuln ID") or f"FINDING-{i+1:04d}"
            )
            f_["title"] = (
                row.get("Title") or row.get("Name") or
                row.get("Vulnerability") or row.get("Finding") or ""
            )
            f_["severity"] = normalize_severity(sev)
            f_["cve"] = row.get("CVE") or row.get("CVE ID") or ""
            f_["asset"] = (
                row.get("Asset") or row.get("Host") or
                row.get("IP") or row.get("IP Address") or
                row.get("Resource") or ""
            )
            f_["description"] = (
                row.get("Description") or row.get("Synopsis") or
                row.get("Details") or row.get("Threat") or ""
            )
            f_["solution"] = (
                row.get("Solution") or row.get("Remediation") or
                row.get("Fix") or row.get("Resolution") or ""
            )
            f_["first_seen"] = (
                row.get("First Seen") or row.get("First Detected") or
                row.get("Discovered") or row.get("Date") or ""
            )
            findings.append(f_)
    return findings


# ---------------------------------------------------------------------------
# Severity normalization
# ---------------------------------------------------------------------------

def normalize_severity(raw):
    r = raw.strip().lower()
    if r in ("critical", "crit"):
        return "Critical"
    if r in ("high", "hi"):
        return "High"
    if r in ("medium", "med", "moderate"):
        return "Medium"
    if r in ("low"):
        return "Low"
    return raw.title()


def normalize_severity_qualys(raw):
    """Qualys uses numeric 1-5 scale."""
    mapping = {"5": "Critical", "4": "High", "3": "Medium", "2": "Low", "1": "Low"}
    return mapping.get(raw.strip(), normalize_severity(raw))


# ---------------------------------------------------------------------------
# POA&M Excel builder
# ---------------------------------------------------------------------------

POAM_COLUMNS = [
    "POA&M ID",
    "Controls",
    "Weakness Name",
    "Weakness Description",
    "Weakness Detector Source",
    "Weakness Source Identifier",
    "Asset Identifier",
    "Point of Contact",
    "Resources Required",
    "Overall Remediation Plan",
    "Original Detection Date",
    "Scheduled Completion Date",
    "Planned Milestones",
    "Milestone Changes",
    "Status Date",
    "Vendor Dependence",
    "Last Vendor Check-in Date",
    "Vendor Dependent Product Name",
    "Original Risk Rating",
    "Adjusted Risk Rating",
    "Risk Adjustment",
    "False Positive",
    "Operational Requirement",
    "Deviation Rationale",
    "Supporting Documents",
    "Comments",
    "Auto-Approve",
    "CVE",
]

SEVERITY_COLORS = {
    "Critical": "FF0000",
    "High": "FF6600",
    "Medium": "FFCC00",
    "Low": "00CC00",
}

HEADER_FILL = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def finding_to_poam_row(finding, index, scanner_name):
    today = datetime.today()
    scheduled = today + timedelta(days=90)

    severity = finding.get("severity", "")
    color = SEVERITY_COLORS.get(severity, "FFFFFF")

    row = {
        "POA&M ID": f"POAM-{index:04d}",
        "Controls": finding.get("controls", ""),
        "Weakness Name": finding.get("title", ""),
        "Weakness Description": finding.get("description", ""),
        "Weakness Detector Source": scanner_name,
        "Weakness Source Identifier": finding.get("plugin_id") or finding.get("finding_id", ""),
        "Asset Identifier": finding.get("asset", ""),
        "Point of Contact": "",
        "Resources Required": "",
        "Overall Remediation Plan": finding.get("solution", ""),
        "Original Detection Date": finding.get("first_seen") or today.strftime("%Y-%m-%d"),
        "Scheduled Completion Date": scheduled.strftime("%Y-%m-%d"),
        "Planned Milestones": "",
        "Milestone Changes": "",
        "Status Date": today.strftime("%Y-%m-%d"),
        "Vendor Dependence": "",
        "Last Vendor Check-in Date": "",
        "Vendor Dependent Product Name": "",
        "Original Risk Rating": severity,
        "Adjusted Risk Rating": severity,
        "Risk Adjustment": "",
        "False Positive": "No",
        "Operational Requirement": "",
        "Deviation Rationale": "",
        "Supporting Documents": "",
        "Comments": "",
        "Auto-Approve": "",
        "CVE": finding.get("cve", ""),
    }
    return row, color


def load_existing_poam(output_path):
    """Load existing POA&M and return (workbook, next_id, existing_dedup_keys)."""
    wb = openpyxl.load_workbook(output_path)
    ws_open = wb["Open POA&M Items"]

    existing_keys = set()
    max_id = 0

    for row in ws_open.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        poam_id = row[POAM_COLUMNS.index("POA&M ID")] or ""
        weakness_name = row[POAM_COLUMNS.index("Weakness Name")] or ""
        asset = row[POAM_COLUMNS.index("Asset Identifier")] or ""
        cve = row[POAM_COLUMNS.index("CVE")] or ""

        # Dedup key: title + asset (CVE may be blank for config findings)
        existing_keys.add((weakness_name.strip().lower(), asset.strip().lower()))
        if cve:
            existing_keys.add((cve.strip().lower(), asset.strip().lower()))

        # Track highest POA&M ID number
        try:
            num = int(poam_id.replace("POAM-", ""))
            max_id = max(max_id, num)
        except (ValueError, AttributeError):
            pass

    return wb, max_id + 1, existing_keys


def build_poam_excel(findings, scanner_name, output_path):
    output_path = Path(output_path)
    existing_keys = set()
    next_id = 1

    if output_path.exists():
        print(f"Existing POA&M found — appending to {output_path}")
        wb, next_id, existing_keys = load_existing_poam(output_path)
        ws_open = wb["Open POA&M Items"]
    else:
        wb = openpyxl.Workbook()
        ws_open = wb.active
        ws_open.title = "Open POA&M Items"

        # Header row
        for col_idx, col_name in enumerate(POAM_COLUMNS, start=1):
            cell = ws_open.cell(row=1, column=col_idx, value=col_name)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER
        ws_open.row_dimensions[1].height = 30

        # Closed sheet
        col_widths = _col_widths()
        ws_closed = wb.create_sheet("Closed POA&M Items")
        for col_idx, col_name in enumerate(POAM_COLUMNS, start=1):
            cell = ws_closed.cell(row=1, column=col_idx, value=col_name)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER
            ws_closed.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 15)
        ws_closed.row_dimensions[1].height = 30
        ws_closed.freeze_panes = "A2"

    col_widths = _col_widths()

    # Filter duplicates
    new_findings = []
    skipped = 0
    for f in findings:
        title_key = (f.get("title", "").strip().lower(), f.get("asset", "").strip().lower())
        cve_key = (f.get("cve", "").strip().lower(), f.get("asset", "").strip().lower())
        if title_key in existing_keys or (f.get("cve") and cve_key in existing_keys):
            skipped += 1
            continue
        new_findings.append(f)

    if skipped:
        print(f"Skipped {skipped} duplicate finding(s) already in POA&M.")

    # Append new rows
    current_row = ws_open.max_row + 1
    for i, finding in enumerate(new_findings):
        row_data, sev_color = finding_to_poam_row(finding, next_id + i, scanner_name)
        sev_fill = PatternFill(start_color=sev_color, end_color=sev_color, fill_type="solid")

        for col_idx, col_name in enumerate(POAM_COLUMNS, start=1):
            cell = ws_open.cell(row=current_row + i, column=col_idx, value=row_data.get(col_name, ""))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER
            if col_name in ("Original Risk Rating", "Adjusted Risk Rating"):
                cell.fill = sev_fill
                cell.font = Font(bold=True, name="Calibri", size=10)

    # Column widths
    for col_idx, col_name in enumerate(POAM_COLUMNS, start=1):
        ws_open.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 15)

    ws_open.freeze_panes = "A2"

    wb.save(output_path)
    print(f"\nPOA&M saved to: {output_path}")
    print(f"New findings added: {len(new_findings)}")
    print(f"Total rows in POA&M: {ws_open.max_row - 1}")

    from collections import Counter
    counts = Counter(f.get("severity", "Unknown") for f in new_findings)
    if counts:
        print("\nNew findings by severity:")
        for sev in ["Critical", "High", "Medium", "Low"]:
            if counts.get(sev):
                print(f"  {sev}: {counts[sev]}")


def _col_widths():
    return {
        "POA&M ID": 12,
        "Controls": 15,
        "Weakness Name": 35,
        "Weakness Description": 50,
        "Weakness Detector Source": 20,
        "Weakness Source Identifier": 22,
        "Asset Identifier": 20,
        "Point of Contact": 20,
        "Resources Required": 20,
        "Overall Remediation Plan": 50,
        "Original Detection Date": 20,
        "Scheduled Completion Date": 22,
        "Planned Milestones": 25,
        "Milestone Changes": 20,
        "Status Date": 15,
        "Vendor Dependence": 18,
        "Last Vendor Check-in Date": 22,
        "Vendor Dependent Product Name": 25,
        "Original Risk Rating": 18,
        "Adjusted Risk Rating": 18,
        "Risk Adjustment": 15,
        "False Positive": 14,
        "Operational Requirement": 22,
        "Deviation Rationale": 22,
        "Supporting Documents": 22,
        "Comments": 30,
        "Auto-Approve": 14,
        "CVE": 18,
    }


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

PARSERS = {
    "nessus": None,  # determined by file extension below
    "tenable": parse_tenable_csv,
    "qualys": parse_qualys_csv,
    "wiz": parse_wiz_csv,
    "generic": parse_generic_csv,
}

SCANNER_LABELS = {
    "nessus": "Nessus",
    "tenable": "Tenable.io",
    "qualys": "Qualys",
    "wiz": "Wiz",
    "generic": "Vulnerability Scanner",
}


def main():
    parser = argparse.ArgumentParser(
        description="Convert scanner exports to a FedRAMP POA&M Excel file."
    )
    parser.add_argument("--input", required=True, help="Path to scanner export file")
    parser.add_argument(
        "--scanner",
        required=True,
        choices=["nessus", "tenable", "qualys", "wiz", "generic"],
        help="Scanner type",
    )
    parser.add_argument("--output", default="poam_output.xlsx", help="Output Excel file path")
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"Error: Input file not found: {args.input}")
        sys.exit(1)

    print(f"Parsing {args.scanner} export: {args.input}")

    if args.scanner == "nessus":
        if input_path.suffix.lower() == ".xml" or input_path.suffix.lower() == ".nessus":
            findings = parse_nessus_xml(args.input)
        else:
            findings = parse_nessus_csv(args.input)
    else:
        parse_fn = PARSERS[args.scanner]
        findings = parse_fn(args.input)

    if not findings:
        print("No findings found (or all were informational). Check your input file.")
        sys.exit(0)

    print(f"Found {len(findings)} findings to add to POA&M.")
    build_poam_excel(findings, SCANNER_LABELS[args.scanner], args.output)


if __name__ == "__main__":
    main()
