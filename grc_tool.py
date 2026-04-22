import argparse
import sys
from datetime import datetime, date, timedelta
from pathlib import Path
from collections import Counter, defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from poam_converter import (
        parse_nessus_csv, parse_nessus_xml,
        parse_tenable_csv, parse_qualys_csv,
        parse_wiz_csv, parse_generic_csv,
        build_poam_excel,
        POAM_COLUMNS,
        HEADER_FILL, HEADER_FONT, THIN_BORDER,
        SEVERITY_COLORS, _col_widths,
    )
except ImportError:
    print("ERROR: poam_converter.py not found. Make sure it is in the same directory.")
    sys.exit(1)


COL = {name: idx + 1 for idx, name in enumerate(POAM_COLUMNS)}

SCANNER_LABELS = {
    "nessus": "Nessus",
    "tenable": "Tenable.io",
    "qualys": "Qualys",
    "wiz": "Wiz",
    "generic": "Vulnerability Scanner",
}

NIST_KEYWORD_MAP = [
    ({"cve-", "patch", "unpatched", "outdated", "end-of-life", "eol",
      "upgrade", "update required", "out of date"},
     ["SI-2", "RA-5"]),
    ({"remote code execution", "rce", "arbitrary code"},
     ["SI-2", "RA-5", "SI-3"]),
    ({"authentication bypass", "auth bypass", "unauthenticated",
      "unauthorized", "privilege escalation", "privilege", "sudo"},
     ["IA-2", "IA-5", "AC-3"]),
    ({"iam", "access key", "permission", "role", "policy", "least privilege",
      "overpermissive", "wildcard", "rbac"},
     ["AC-2", "AC-3", "AC-6"]),
    ({"mfa", "multi-factor", "two-factor", "2fa", "totp"},
     ["IA-2", "IA-5"]),
    ({"password", "credential", "default password", "weak password",
      "hardcoded", "cleartext password", "plaintext"},
     ["IA-5", "IA-6"]),
    ({"unencrypted", "encryption at rest", "not encrypted", "ebs",
      "disk encryption"},
     ["SC-28"]),
    ({"ssl", "tls", "weak cipher", "weak algorithm", "weak hash",
      "sha1", "md5", "cleartext", "plaintext traffic"},
     ["SC-8", "SC-23"]),
    ({"certificate", "cert", "expired cert", "self-signed"},
     ["SC-8", "IA-5"]),
    ({"security group", "firewall", "inbound", "unrestricted",
      "0.0.0.0", "open port", "exposed port", "network access"},
     ["SC-7", "AC-17"]),
    ({"ssh", "secure shell"},
     ["AC-17", "SC-8", "IA-2"]),
    ({"vpn", "remote access", "rdp", "remote desktop"},
     ["AC-17", "IA-2"]),
    ({"logging", "audit", "cloudtrail", "siem",
      "audit trail", "event log", "monitoring disabled"},
     ["AU-2", "AU-12"]),
    ({"log integrity", "log tamper", "log validation"},
     ["AU-9"]),
    ({"misconfiguration", "misconfigured", "default config",
      "hardening", "benchmark", "cis", "stig", "configuration"},
     ["CM-6", "CM-7"]),
    ({"unnecessary service", "unused service", "trace method",
      "expn", "debug enabled", "test endpoint"},
     ["CM-7"]),
    ({"supply chain", "third-party", "dependency", "library",
      "component", "npm", "pip", "maven"},
     ["SA-12", "SI-7"]),
    ({"malware", "antivirus", "anti-virus", "endpoint protection"},
     ["SI-3"]),
    ({"container", "docker", "kubernetes", "k8s", "pod",
      "running as root", "privileged container"},
     ["CM-6", "CM-7", "AC-6"]),
    ({"s3", "bucket", "blob storage", "object storage", "public bucket",
      "storage account"},
     ["AC-3", "SC-28", "AC-6"]),
    ({"secret", "api key", "token", "private key", "key rotation",
      "secrets manager", "vault"},
     ["SC-12", "IA-5"]),
    ({"sql injection", "sqli", "xss", "cross-site", "injection",
      "input validation", "command injection"},
     ["SI-10", "SI-3"]),
    ({"file disclosure", "directory traversal", "path traversal",
      "arbitrary file read", "lfi", "rfi"},
     ["AC-3", "SI-10"]),
    ({"backup", "recovery", "restore", "disaster recovery"},
     ["CP-9", "CP-10"]),
    ({"ntp", "time sync", "clock"},
     ["AU-8"]),
    ({"dns", "domain hijack", "subdomain takeover"},
     ["SC-20", "SC-21"]),
    ({"denial of service", "ddos", "resource exhaustion"},
     ["SC-5"]),
    ({"asset inventory", "unmanaged asset", "rogue device"},
     ["CM-8"]),
    ({"smb", "print spooler", "printnightmare", "smbghost"},
     ["CM-6", "SI-2"]),
]


def load_poam(path):
    p = Path(path)
    if not p.exists():
        print(f"ERROR: POA&M file not found: {path}")
        sys.exit(1)
    wb = openpyxl.load_workbook(path)
    if "Open POA&M Items" not in wb.sheetnames:
        print(f"ERROR: Sheet 'Open POA&M Items' not found in {path}")
        sys.exit(1)
    if "Closed POA&M Items" not in wb.sheetnames:
        print(f"ERROR: Sheet 'Closed POA&M Items' not found in {path}")
        sys.exit(1)
    return wb, wb["Open POA&M Items"], wb["Closed POA&M Items"]


def save_poam(wb, path):
    try:
        wb.save(path)
    except PermissionError:
        print(f"ERROR: Cannot save — is {path} open in Excel? Close it and try again.")
        sys.exit(1)


def get_all_rows(ws):
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        rows.append(dict(zip(POAM_COLUMNS, row)))
    return rows


def find_row_by_id(ws, poam_id):
    target = poam_id.upper().strip()
    for row_num in range(2, ws.max_row + 1):
        val = ws.cell(row=row_num, column=COL["POA&M ID"]).value
        if val and str(val).upper().strip() == target:
            row_dict = {}
            for col_name in POAM_COLUMNS:
                row_dict[col_name] = ws.cell(row=row_num, column=COL[col_name]).value
            return row_num, row_dict
    return None, None


def normalize_id(raw_id):
    raw = raw_id.strip().upper()
    if raw.startswith("POAM-"):
        return raw
    try:
        return f"POAM-{int(raw):04d}"
    except ValueError:
        return raw


def map_controls(title, description, cve):
    text = f"{title} {description} {cve}".lower()
    controls = []
    for keywords, ctrl_list in NIST_KEYWORD_MAP:
        if any(kw in text for kw in keywords):
            controls.extend(ctrl_list)
    if cve and cve.upper().startswith("CVE-"):
        controls.extend(["SI-2", "RA-5"])
    return ", ".join(sorted(set(controls))) if controls else ""


def parse_date(val):
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.date() if isinstance(val, datetime) else val
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%Y/%m/%d", "%m-%d-%Y"):
        try:
            return datetime.strptime(str(val).strip(), fmt).date()
        except ValueError:
            continue
    return None


def is_overdue(date_val):
    d = parse_date(date_val)
    if d is None:
        return False
    return d < date.today()


def days_overdue(date_val):
    d = parse_date(date_val)
    if d is None:
        return 0
    delta = (date.today() - d).days
    return max(delta, 0)


def call_ollama(prompt):
    try:
        import requests
        resp = requests.post(
            "http://localhost:11434/api/generate",
            json={"model": "mistral", "prompt": prompt, "stream": False},
            timeout=30,
        )
        resp.raise_for_status()
        return resp.json().get("response", "").strip()
    except Exception:
        return None


def apply_row_style(ws, row_num, severity):
    color = SEVERITY_COLORS.get(severity, "FFFFFF")
    sev_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    for col_idx in range(1, len(POAM_COLUMNS) + 1):
        cell = ws.cell(row=row_num, column=col_idx)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if POAM_COLUMNS[col_idx - 1] in ("Original Risk Rating", "Adjusted Risk Rating"):
            cell.fill = sev_fill
            cell.font = Font(bold=True, name="Calibri", size=10)


def cmd_convert(args):
    input_path = Path(args.input)
    if not input_path.exists():
        print(f"ERROR: Input file not found: {args.input}")
        sys.exit(1)

    print(f"Parsing {args.scanner} export: {args.input}")

    if args.scanner == "nessus":
        if input_path.suffix.lower() in (".xml", ".nessus"):
            findings = parse_nessus_xml(args.input)
        else:
            findings = parse_nessus_csv(args.input)
    elif args.scanner == "tenable":
        findings = parse_tenable_csv(args.input)
    elif args.scanner == "qualys":
        findings = parse_qualys_csv(args.input)
    elif args.scanner == "wiz":
        findings = parse_wiz_csv(args.input)
    else:
        findings = parse_generic_csv(args.input)

    if not findings:
        print("No findings found (or all were informational). Check your input file.")
        sys.exit(0)

    print(f"Found {len(findings)} findings to add to POA&M.")
    build_poam_excel(findings, SCANNER_LABELS[args.scanner], args.output)


def cmd_enrich(args):
    wb, ws_open, _ = load_poam(args.poam)
    today_str = date.today().strftime("%Y-%m-%d")

    controls_added = 0
    ai_enriched = 0
    skipped_controls = 0
    skipped_ai = 0
    ollama_unavailable = False

    print(f"Enriching {args.poam}...")
    if args.ai:
        print("AI mode enabled — connecting to Ollama/Mistral...")

    for row_num in range(2, ws_open.max_row + 1):
        poam_id = ws_open.cell(row=row_num, column=COL["POA&M ID"]).value
        if not poam_id:
            continue

        title = ws_open.cell(row=row_num, column=COL["Weakness Name"]).value or ""
        desc = ws_open.cell(row=row_num, column=COL["Weakness Description"]).value or ""
        cve = ws_open.cell(row=row_num, column=COL["CVE"]).value or ""
        existing_controls = ws_open.cell(row=row_num, column=COL["Controls"]).value

        if not existing_controls:
            mapped = map_controls(title, desc, cve)
            if mapped:
                ws_open.cell(row=row_num, column=COL["Controls"]).value = mapped
                controls_added += 1
        else:
            skipped_controls += 1

        if args.ai and not ollama_unavailable:
            existing_remediation = ws_open.cell(row=row_num, column=COL["Overall Remediation Plan"]).value
            if not existing_remediation:
                prompt = (
                    f"You are a FedRAMP GRC analyst. Write a 2-3 sentence remediation plan "
                    f"for the following vulnerability finding. Be specific and actionable.\n\n"
                    f"Finding: {title}\n"
                    f"Description: {desc}\n"
                    f"CVE: {cve if cve else 'N/A'}\n\n"
                    f"Remediation Plan:"
                )
                result = call_ollama(prompt)
                if result is None:
                    print("[AI] Ollama not available — skipping AI enrichment for remaining findings.")
                    ollama_unavailable = True
                else:
                    ws_open.cell(row=row_num, column=COL["Overall Remediation Plan"]).value = result
                    ai_enriched += 1
            else:
                skipped_ai += 1

    save_poam(wb, args.poam)
    print(f"\nEnrichment complete:")
    print(f"  NIST controls added:     {controls_added}")
    print(f"  NIST controls skipped (already set): {skipped_controls}")
    if args.ai:
        print(f"  AI remediation added:    {ai_enriched}")
        print(f"  AI remediation skipped (already set): {skipped_ai}")


def cmd_close(args):
    poam_id = normalize_id(args.poam_id)
    wb, ws_open, ws_closed = load_poam(args.poam)

    row_num, row_dict = find_row_by_id(ws_open, poam_id)
    if row_num is None:
        print(f"ERROR: Finding '{poam_id}' not found in Open POA&M Items.")
        sys.exit(1)

    severity = row_dict.get("Original Risk Rating", "")
    closure_note = f"Closed: {args.method} on {args.close_date}"

    closed_row = ws_closed.max_row + 1
    if closed_row == 1:
        closed_row = 2

    for col_name in POAM_COLUMNS:
        val = row_dict.get(col_name, "")
        ws_closed.cell(row=closed_row, column=COL[col_name]).value = val

    ws_closed.cell(row=closed_row, column=COL["Status Date"]).value = args.close_date
    ws_closed.cell(row=closed_row, column=COL["Comments"]).value = closure_note

    apply_row_style(ws_closed, closed_row, severity)

    ws_open.delete_rows(row_num)

    save_poam(wb, args.poam)
    print(f"Finding {poam_id} moved to Closed POA&M Items.")
    print(f"  Method: {args.method}")
    print(f"  Date:   {args.close_date}")


def cmd_update(args):
    poam_id = normalize_id(args.poam_id)
    wb, ws_open, _ = load_poam(args.poam)

    row_num, row_dict = find_row_by_id(ws_open, poam_id)
    if row_num is None:
        print(f"ERROR: Finding '{poam_id}' not found in Open POA&M Items.")
        sys.exit(1)

    updated = []
    today_str = date.today().strftime("%Y-%m-%d")

    if args.milestone:
        ws_open.cell(row=row_num, column=COL["Milestone Changes"]).value = args.milestone
        updated.append(f"Milestone Changes = {args.milestone}")

    if args.poc:
        ws_open.cell(row=row_num, column=COL["Point of Contact"]).value = args.poc
        updated.append(f"Point of Contact = {args.poc}")

    if args.vendor_date:
        ws_open.cell(row=row_num, column=COL["Last Vendor Check-in Date"]).value = args.vendor_date
        updated.append(f"Last Vendor Check-in Date = {args.vendor_date}")

    if args.status:
        existing = ws_open.cell(row=row_num, column=COL["Comments"]).value or ""
        status_entry = f"[STATUS {today_str}: {args.status}]"
        ws_open.cell(row=row_num, column=COL["Comments"]).value = f"{status_entry} {existing}".strip()
        updated.append(f"Status = {args.status}")

    ws_open.cell(row=row_num, column=COL["Status Date"]).value = today_str

    save_poam(wb, args.poam)

    if updated:
        print(f"Updated {poam_id}:")
        for u in updated:
            print(f"  {u}")
    else:
        print(f"No fields specified to update for {poam_id}. Use --milestone, --poc, --vendor-date, or --status.")


def cmd_deviation(args):
    poam_id = normalize_id(args.poam_id)
    wb, ws_open, _ = load_poam(args.poam)

    row_num, row_dict = find_row_by_id(ws_open, poam_id)
    if row_num is None:
        print(f"ERROR: Finding '{poam_id}' not found in Open POA&M Items.")
        sys.exit(1)

    today_str = date.today().strftime("%Y-%m-%d")

    if args.dev_type == "fp":
        ws_open.cell(row=row_num, column=COL["False Positive"]).value = "Yes"
        ws_open.cell(row=row_num, column=COL["Operational Requirement"]).value = ""
        label = "False Positive"
    else:
        ws_open.cell(row=row_num, column=COL["Operational Requirement"]).value = "Yes"
        ws_open.cell(row=row_num, column=COL["False Positive"]).value = ""
        label = "Operational Requirement"

    ws_open.cell(row=row_num, column=COL["Deviation Rationale"]).value = args.rationale
    ws_open.cell(row=row_num, column=COL["Status Date"]).value = today_str

    save_poam(wb, args.poam)
    print(f"Finding {poam_id} marked as {label}.")
    print(f"  Rationale: {args.rationale}")


def cmd_dashboard(args):
    wb, ws_open, ws_closed = load_poam(args.poam)
    open_rows = get_all_rows(ws_open)
    closed_rows = get_all_rows(ws_closed)

    today = date.today()
    overdue = [r for r in open_rows if is_overdue(r.get("Scheduled Completion Date"))]
    due_this_month = [
        r for r in open_rows
        if not is_overdue(r.get("Scheduled Completion Date")) and
        (d := parse_date(r.get("Scheduled Completion Date"))) and
        d.year == today.year and d.month == today.month
    ]

    sev_counts = Counter(r.get("Original Risk Rating", "Unknown") for r in open_rows)

    dated = [(r, parse_date(r.get("Original Detection Date"))) for r in open_rows]
    dated = [(r, d) for r, d in dated if d is not None]
    dated.sort(key=lambda x: x[1])
    top5 = dated[:5]

    w = 62
    print("=" * w)
    print(f"  GRC DASHBOARD — {Path(args.poam).name}")
    print(f"  Generated: {today.strftime('%Y-%m-%d')}")
    print("=" * w)
    print(f"  {'OPEN ITEMS:':<22}{len(open_rows):<10}{'CLOSED ITEMS:':<18}{len(closed_rows)}")
    print(f"  {'OVERDUE:':<22}{len(overdue):<10}{'DUE THIS MONTH:':<18}{len(due_this_month)}")
    print("-" * w)
    print("  SEVERITY BREAKDOWN (Open)")
    for sev in ["Critical", "High", "Medium", "Low"]:
        count = sev_counts.get(sev, 0)
        bar = "#" * count
        print(f"    {sev:<12} {count:>4}  {bar}")
    print("-" * w)
    print("  TOP 5 OLDEST OPEN FINDINGS")
    if top5:
        for i, (r, d) in enumerate(top5, 1):
            poam_id = r.get("POA&M ID", "?")
            sev = r.get("Original Risk Rating", "?")
            name = (r.get("Weakness Name") or "")[:38]
            scheduled = r.get("Scheduled Completion Date")
            overdue_days = days_overdue(scheduled)
            overdue_str = f"({overdue_days}d overdue)" if overdue_days > 0 else "(on track)"
            print(f"    {i}. {poam_id}  [{sev:<8}]  {name:<40} {overdue_str}")
    else:
        print("    No open findings.")
    print("=" * w)


def cmd_report(args):
    wb_src, ws_open, ws_closed = load_poam(args.poam)
    open_rows = get_all_rows(ws_open)
    closed_rows = get_all_rows(ws_closed)

    today = date.today()
    output = args.output or f"grc_report_{today.strftime('%Y%m%d')}.xlsx"

    wb = openpyxl.Workbook()
    col_widths = _col_widths()

    ws_sum = wb.active
    ws_sum.title = "Summary"

    def hcell(ws, row, col, value):
        cell = ws.cell(row=row, column=col, value=value)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        return cell

    def dcell(ws, row, col, value, bold=False, fill_color=None):
        cell = ws.cell(row=row, column=col, value=value)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        if bold:
            cell.font = Font(bold=True, name="Calibri", size=11)
        if fill_color:
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        return cell

    ws_sum.row_dimensions[1].height = 24
    for col, header in enumerate(["Severity", "Open", "Closed", "Total"], start=1):
        hcell(ws_sum, 1, col, header)
        ws_sum.column_dimensions[get_column_letter(col)].width = 18

    open_sev = Counter(r.get("Original Risk Rating", "Unknown") for r in open_rows)
    closed_sev = Counter(r.get("Original Risk Rating", "Unknown") for r in closed_rows)

    row = 2
    for sev in ["Critical", "High", "Medium", "Low"]:
        color = SEVERITY_COLORS.get(sev, "FFFFFF")
        dcell(ws_sum, row, 1, sev, fill_color=color)
        dcell(ws_sum, row, 2, open_sev.get(sev, 0))
        dcell(ws_sum, row, 3, closed_sev.get(sev, 0))
        dcell(ws_sum, row, 4, open_sev.get(sev, 0) + closed_sev.get(sev, 0))
        row += 1

    dcell(ws_sum, row, 1, "TOTAL", bold=True)
    dcell(ws_sum, row, 2, len(open_rows), bold=True)
    dcell(ws_sum, row, 3, len(closed_rows), bold=True)
    dcell(ws_sum, row, 4, len(open_rows) + len(closed_rows), bold=True)
    row += 2

    overdue_count = sum(1 for r in open_rows if is_overdue(r.get("Scheduled Completion Date")))
    overdue_pct = round(overdue_count / len(open_rows) * 100, 1) if open_rows else 0
    ws_sum.cell(row=row, column=1, value="Overdue Findings").font = Font(bold=True, name="Calibri")
    ws_sum.cell(row=row, column=2, value=overdue_count)
    ws_sum.cell(row=row, column=3, value=f"{overdue_pct}% of open")

    ws_age = wb.create_sheet("Aging")
    age_headers = ["POA&M ID", "Weakness Name", "Asset Identifier", "Original Risk Rating",
                   "Original Detection Date", "Scheduled Completion Date", "Days Overdue"]
    for col, h in enumerate(age_headers, start=1):
        cell = hcell(ws_age, 1, col, h)
        ws_age.column_dimensions[get_column_letter(col)].width = [12, 40, 20, 18, 20, 22, 14][col - 1]
    ws_age.row_dimensions[1].height = 24
    ws_age.freeze_panes = "A2"

    aging_data = [
        (r, days_overdue(r.get("Scheduled Completion Date")))
        for r in open_rows
    ]
    aging_data.sort(key=lambda x: x[1], reverse=True)

    for row_idx, (r, d_over) in enumerate(aging_data, start=2):
        vals = [
            r.get("POA&M ID"), r.get("Weakness Name"), r.get("Asset Identifier"),
            r.get("Original Risk Rating"), r.get("Original Detection Date"),
            r.get("Scheduled Completion Date"), d_over
        ]
        for col, val in enumerate(vals, start=1):
            cell = ws_age.cell(row=row_idx, column=col, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        d_cell = ws_age.cell(row=row_idx, column=7)
        if d_over > 90:
            d_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            d_cell.font = Font(bold=True, color="FFFFFF")
        elif d_over > 30:
            d_cell.fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
        elif d_over > 0:
            d_cell.fill = PatternFill(start_color="FFCC00", end_color="FFCC00", fill_type="solid")
        else:
            d_cell.fill = PatternFill(start_color="00CC00", end_color="00CC00", fill_type="solid")

    ws_scan = wb.create_sheet("By Scanner")
    scan_headers = ["Scanner", "POA&M ID", "Weakness Name", "Original Risk Rating", "Scheduled Completion Date"]
    for col, h in enumerate(scan_headers, start=1):
        hcell(ws_scan, 1, col, h)
        ws_scan.column_dimensions[get_column_letter(col)].width = [20, 12, 45, 18, 22][col - 1]
    ws_scan.row_dimensions[1].height = 24
    ws_scan.freeze_panes = "A2"

    sev_order = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3}
    scanner_rows = sorted(
        open_rows,
        key=lambda r: (
            r.get("Weakness Detector Source", ""),
            sev_order.get(r.get("Original Risk Rating", ""), 99)
        )
    )

    for row_idx, r in enumerate(scanner_rows, start=2):
        for col, key in enumerate(["Weakness Detector Source", "POA&M ID", "Weakness Name",
                                    "Original Risk Rating", "Scheduled Completion Date"], start=1):
            cell = ws_scan.cell(row=row_idx, column=col, value=r.get(key, ""))
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    try:
        wb.save(output)
        print(f"Executive report saved to: {output}")
        print(f"  Open findings:   {len(open_rows)}")
        print(f"  Closed findings: {len(closed_rows)}")
        print(f"  Overdue:         {overdue_count}")
    except PermissionError:
        print(f"ERROR: Cannot save — is {output} open in Excel?")
        sys.exit(1)


def cmd_conmon(args):
    wb_src, ws_open, ws_closed = load_poam(args.poam)
    open_rows = get_all_rows(ws_open)
    closed_rows = get_all_rows(ws_closed)

    today = date.today()
    if args.month:
        try:
            year, month = args.month.strip().split("-")
            year, month = int(year), int(month)
        except ValueError:
            print("ERROR: --month must be in format YYYY-MM (e.g. 2026-04)")
            sys.exit(1)
    else:
        year, month = today.year, today.month

    month_prefix = f"{year}-{month:02d}"
    output = args.output or f"conmon_{year}_{month:02d}.xlsx"

    def in_month(date_val):
        if date_val is None:
            return False
        return str(date_val).startswith(month_prefix)

    opened = [r for r in open_rows if in_month(r.get("Original Detection Date"))]
    closed = [r for r in closed_rows if in_month(r.get("Status Date"))]

    wb = openpyxl.Workbook()
    col_widths = _col_widths()

    def build_sheet(ws, title_text, rows):
        ws.title = title_text
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(POAM_COLUMNS))
        cover = ws.cell(row=1, column=1,
                        value=f"ConMon Report — Month: {month_prefix} | Generated: {today} | System: {Path(args.poam).name}")
        cover.font = Font(bold=True, name="Calibri", size=11)
        cover.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        cover.font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
        cover.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 24

        for col_idx, col_name in enumerate(POAM_COLUMNS, start=1):
            cell = ws.cell(row=2, column=col_idx, value=col_name)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER
            ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 15)
        ws.row_dimensions[2].height = 30
        ws.freeze_panes = "A3"

        for row_idx, r in enumerate(rows, start=3):
            sev = r.get("Original Risk Rating", "")
            color = SEVERITY_COLORS.get(sev, "FFFFFF")
            sev_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            for col_idx, col_name in enumerate(POAM_COLUMNS, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=r.get(col_name, ""))
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if col_name in ("Original Risk Rating", "Adjusted Risk Rating"):
                    cell.fill = sev_fill
                    cell.font = Font(bold=True, name="Calibri", size=10)

    ws_opened = wb.active
    build_sheet(ws_opened, "Opened This Month", opened)
    ws_closed_sheet = wb.create_sheet("Closed This Month")
    build_sheet(ws_closed_sheet, "Closed This Month", closed)

    try:
        wb.save(output)
        print(f"ConMon report saved to: {output}")
        print(f"  Month:          {month_prefix}")
        print(f"  Opened:         {len(opened)}")
        print(f"  Closed:         {len(closed)}")
    except PermissionError:
        print(f"ERROR: Cannot save — is {output} open in Excel?")
        sys.exit(1)


def cmd_export(args):
    wb_src, ws_open, ws_closed = load_poam(args.poam)
    today = date.today()
    output = args.output or f"fedramp_poam_{today.strftime('%Y%m%d')}.xlsx"
    col_widths = _col_widths()

    wb = openpyxl.Workbook()

    def copy_sheet(ws_src, ws_dst):
        for col_idx, col_name in enumerate(POAM_COLUMNS, start=1):
            cell = ws_dst.cell(row=1, column=col_idx, value=col_name)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER
            ws_dst.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 15)
        ws_dst.row_dimensions[1].height = 30
        ws_dst.freeze_panes = "A2"

        dst_row = 2
        for src_row in range(2, ws_src.max_row + 1):
            if not any(ws_src.cell(row=src_row, column=c).value for c in range(1, len(POAM_COLUMNS) + 1)):
                continue
            sev = ws_src.cell(row=src_row, column=COL["Original Risk Rating"]).value or ""
            color = SEVERITY_COLORS.get(sev, "FFFFFF")
            sev_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            for col_idx, col_name in enumerate(POAM_COLUMNS, start=1):
                val = ws_src.cell(row=src_row, column=col_idx).value
                cell = ws_dst.cell(row=dst_row, column=col_idx, value=val)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if col_name in ("Original Risk Rating", "Adjusted Risk Rating"):
                    cell.fill = sev_fill
                    cell.font = Font(bold=True, name="Calibri", size=10)
            dst_row += 1

    ws_open_dst = wb.active
    ws_open_dst.title = "Open POA&M Items"
    copy_sheet(ws_open, ws_open_dst)

    ws_closed_dst = wb.create_sheet("Closed POA&M Items")
    copy_sheet(ws_closed, ws_closed_dst)

    try:
        wb.save(output)
        open_count = ws_open_dst.max_row - 1
        closed_count = ws_closed_dst.max_row - 1
        print(f"FedRAMP POA&M export saved to: {output}")
        print(f"  Open findings:   {open_count}")
        print(f"  Closed findings: {closed_count}")
    except PermissionError:
        print(f"ERROR: Cannot save — is {output} open in Excel?")
        sys.exit(1)


SEVERITY_TO_N_RATING = {
    "Critical": "N4",
    "High":     "N3",
    "Medium":   "N2",
    "Low":      "N1",
}

N_RATING_ORDER = ["N1", "N2", "N3", "N4", "N5"]

VDR_EVAL_DEADLINE_DAYS = {"low": 7, "moderate": 5, "high": 2}

VDR_UPDATE_FREQ_DAYS = {"low": 30, "moderate": 14, "high": 7}

VDR_ACCEPTANCE_THRESHOLD_DAYS = 192

IRV_KEYWORDS = {
    "public", "internet", "external", "exposed", "inbound", "0.0.0.0",
    "unrestricted", "open port", "s3", "bucket", "api gateway", "load balancer",
    "web server", "http", "https", "cdn", "dns", "cloudfront", "publicly accessible",
}

VDR_ACTIVE_COLUMNS = [
    "Tracking ID",
    "Detection Date",
    "Detection Source",
    "Evaluation Deadline",
    "Evaluation Status",
    "Internet Reachability",
    "Likely Exploitability",
    "CISA KEV",
    "EPSS Score",
    "Adverse Impact (N1-N5)",
    "CVE",
    "Weakness Name",
    "Asset Identifier",
    "Current Status",
    "Next Milestone",
    "Scheduled Completion Date",
    "Overdue Explanation",
    "Supplementary Risk Info",
    "Disposition",
]

VDR_ACCEPTED_COLUMNS = [
    "Tracking ID",
    "Detection Date",
    "Detection Source",
    "Evaluation Deadline",
    "Internet Reachability",
    "Likely Exploitability",
    "CISA KEV",
    "EPSS Score",
    "Current Adverse Impact (N1-N5)",
    "CVE",
    "Weakness Name",
    "Asset Identifier",
    "Acceptance Rationale",
    "Agency Mitigation Guidance",
]

N_COLORS = {
    "N5": "FF0000",
    "N4": "FF4500",
    "N3": "FF6600",
    "N2": "FFCC00",
    "N1": "00CC00",
}


def _fetch_cisa_kev():
    try:
        import requests
        resp = requests.get(
            "https://www.cisa.gov/sites/default/files/feeds/known_exploited_vulnerabilities.json",
            timeout=15,
        )
        resp.raise_for_status()
        data = resp.json()
        return {v["cveID"].upper() for v in data.get("vulnerabilities", [])}
    except Exception:
        return set()


def _fetch_epss_scores(cve_list):
    scores = {}
    cves = [c.upper() for c in cve_list if c and c.upper().startswith("CVE-")]
    if not cves:
        return scores
    try:
        import requests
        for i in range(0, len(cves), 100):
            batch = cves[i:i + 100]
            params = {"cve": ",".join(batch)}
            resp = requests.get("https://api.first.org/data/v1/epss", params=params, timeout=15)
            resp.raise_for_status()
            for entry in resp.json().get("data", []):
                scores[entry["cve"].upper()] = round(float(entry["epss"]), 4)
    except Exception:
        pass
    return scores


def _derive_irv(title, description):
    text = f"{title} {description}".lower()
    return "IRV" if any(kw in text for kw in IRV_KEYWORDS) else "NIRV"


def _derive_lev(severity, cve, kev_set, epss_scores):
    cve_upper = (cve or "").upper()
    if cve_upper and cve_upper in kev_set:
        return "LEV"
    if cve_upper and epss_scores.get(cve_upper, 0) >= 0.10:
        return "LEV"
    if severity in ("Critical", "High") and cve:
        return "LEV"
    return "NLEV"


def _derive_n_rating(severity, irv, lev, in_kev):
    base = SEVERITY_TO_N_RATING.get(severity, "N2")
    idx = N_RATING_ORDER.index(base)
    if in_kev:
        idx = max(idx, N_RATING_ORDER.index("N3"))
    if irv == "IRV" and lev == "LEV":
        idx = min(idx + 1, len(N_RATING_ORDER) - 1)
    return N_RATING_ORDER[idx]


def _eval_deadline(detection_date, baseline):
    d = parse_date(detection_date)
    if d is None:
        return ""
    deadline = d + timedelta(days=VDR_EVAL_DEADLINE_DAYS.get(baseline, 5))
    return deadline.strftime("%Y-%m-%d")


def _eval_status(eval_deadline_str):
    d = parse_date(eval_deadline_str)
    if d is None:
        return ""
    return "Overdue" if date.today() > d else "On Track"


def _build_vdr_records(open_rows, baseline, kev_set, epss_scores):
    today = date.today()
    active = []
    accepted = []

    for r in open_rows:
        detection_date = parse_date(r.get("Original Detection Date"))
        age_days = (today - detection_date).days if detection_date else 0
        is_fp = str(r.get("False Positive") or "").strip().lower() == "yes"
        is_or = str(r.get("Operational Requirement") or "").strip().lower() == "yes"

        sev = r.get("Original Risk Rating", "")
        title = r.get("Weakness Name") or ""
        desc = r.get("Weakness Description") or ""
        cve = (r.get("CVE") or "").strip()
        cve_upper = cve.upper()

        irv = _derive_irv(title, desc)
        in_kev = bool(cve_upper and cve_upper in kev_set)
        lev = _derive_lev(sev, cve, kev_set, epss_scores)
        n_rating = _derive_n_rating(sev, irv, lev, in_kev)
        epss = epss_scores.get(cve_upper, "") if cve_upper else ""
        eval_deadline = _eval_deadline(r.get("Original Detection Date"), baseline)
        eval_status = _eval_status(eval_deadline)
        overdue_days = days_overdue(r.get("Scheduled Completion Date"))
        overdue_explanation = (
            f"Remediation is {overdue_days} days past scheduled completion date."
            if overdue_days > 0 else ""
        )

        record = {
            **r,
            "_irv": irv,
            "_lev": lev,
            "_n_rating": n_rating,
            "_in_kev": in_kev,
            "_epss": epss,
            "_eval_deadline": eval_deadline,
            "_eval_status": eval_status,
            "_overdue_explanation": overdue_explanation,
            "_age_days": age_days,
        }

        if age_days >= VDR_ACCEPTANCE_THRESHOLD_DAYS or is_fp or is_or:
            accepted.append(record)
        else:
            active.append(record)

    return active, accepted


def _write_vdr_header(ws, columns):
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = 20
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


def _style_vdr_row(ws, row_num, n_rating, columns, eval_status=None):
    color = N_COLORS.get(n_rating, "FFFFFF")
    n_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    kev_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=row_num, column=col_idx)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if col_name in ("Adverse Impact (N1-N5)", "Current Adverse Impact (N1-N5)"):
            cell.fill = n_fill
            cell.font = Font(bold=True, name="Calibri", size=10)
        elif col_name == "Evaluation Status" and eval_status == "Overdue":
            cell.fill = red_fill
            cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        elif col_name == "CISA KEV" and cell.value == "Yes":
            cell.fill = kev_fill
            cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)


def cmd_vdr(args):
    wb_src, ws_open, _ = load_poam(args.poam)
    open_rows = get_all_rows(ws_open)
    today = date.today()
    baseline = args.baseline
    output_xlsx = args.output or f"vdr_{today.strftime('%Y%m%d')}.xlsx"
    output_json = output_xlsx.replace(".xlsx", ".json")

    all_cves = [r.get("CVE", "") for r in open_rows if r.get("CVE")]

    print(f"Fetching CISA KEV catalog...")
    kev_set = _fetch_cisa_kev()
    if kev_set:
        print(f"  Loaded {len(kev_set):,} known exploited vulnerabilities.")
    else:
        print("  CISA KEV unavailable (offline) — LEV classification will use severity fallback.")

    print(f"Fetching EPSS scores for {len(all_cves)} CVEs...")
    epss_scores = _fetch_epss_scores(all_cves)
    if epss_scores:
        print(f"  Retrieved EPSS scores for {len(epss_scores)} CVEs.")
    else:
        print("  EPSS unavailable (offline) — LEV classification will use severity fallback.")

    active, accepted = _build_vdr_records(open_rows, baseline, kev_set, epss_scores)

    kev_count = sum(1 for r in active + accepted if r["_in_kev"])
    eval_overdue = sum(1 for r in active if r["_eval_status"] == "Overdue")

    wb = openpyxl.Workbook()

    ws_active = wb.active
    ws_active.title = "Active Vulnerabilities"
    _write_vdr_header(ws_active, VDR_ACTIVE_COLUMNS)

    for row_idx, r in enumerate(active, start=2):
        scheduled = r.get("Scheduled Completion Date") or ""
        row_data = {
            "Tracking ID":              r.get("POA&M ID", ""),
            "Detection Date":           r.get("Original Detection Date", ""),
            "Detection Source":         r.get("Weakness Detector Source", ""),
            "Evaluation Deadline":      r["_eval_deadline"],
            "Evaluation Status":        r["_eval_status"],
            "Internet Reachability":    r["_irv"],
            "Likely Exploitability":    r["_lev"],
            "CISA KEV":                 "Yes" if r["_in_kev"] else "No",
            "EPSS Score":               r["_epss"],
            "Adverse Impact (N1-N5)":   r["_n_rating"],
            "CVE":                      r.get("CVE", ""),
            "Weakness Name":            r.get("Weakness Name", ""),
            "Asset Identifier":         r.get("Asset Identifier", ""),
            "Current Status":           r.get("Comments", "") or "Open",
            "Next Milestone":           r.get("Milestone Changes", "") or str(scheduled),
            "Scheduled Completion Date": scheduled,
            "Overdue Explanation":      r["_overdue_explanation"],
            "Supplementary Risk Info":  r.get("Overall Remediation Plan", ""),
            "Disposition":              "Active",
        }
        for col_idx, col_name in enumerate(VDR_ACTIVE_COLUMNS, start=1):
            ws_active.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ""))
        _style_vdr_row(ws_active, row_idx, r["_n_rating"], VDR_ACTIVE_COLUMNS, r["_eval_status"])

    ws_acc = wb.create_sheet("Accepted Vulnerabilities")
    _write_vdr_header(ws_acc, VDR_ACCEPTED_COLUMNS)

    for row_idx, r in enumerate(accepted, start=2):
        is_fp = str(r.get("False Positive") or "").strip().lower() == "yes"
        is_or = str(r.get("Operational Requirement") or "").strip().lower() == "yes"
        age_days = r["_age_days"]

        if is_fp:
            rationale = f"False Positive — {r.get('Deviation Rationale', '')}"
        elif is_or:
            rationale = f"Operational Requirement — {r.get('Deviation Rationale', '')}"
        else:
            rationale = (
                f"Finding exceeded {VDR_ACCEPTANCE_THRESHOLD_DAYS}-day remediation threshold "
                f"({age_days} days since detection). {r.get('Deviation Rationale', '')}"
            )

        row_data = {
            "Tracking ID":                    r.get("POA&M ID", ""),
            "Detection Date":                 r.get("Original Detection Date", ""),
            "Detection Source":               r.get("Weakness Detector Source", ""),
            "Evaluation Deadline":            r["_eval_deadline"],
            "Internet Reachability":          r["_irv"],
            "Likely Exploitability":          r["_lev"],
            "CISA KEV":                       "Yes" if r["_in_kev"] else "No",
            "EPSS Score":                     r["_epss"],
            "Current Adverse Impact (N1-N5)": r["_n_rating"],
            "CVE":                            r.get("CVE", ""),
            "Weakness Name":                  r.get("Weakness Name", ""),
            "Asset Identifier":               r.get("Asset Identifier", ""),
            "Acceptance Rationale":           rationale.strip(),
            "Agency Mitigation Guidance":     r.get("Overall Remediation Plan", ""),
        }
        for col_idx, col_name in enumerate(VDR_ACCEPTED_COLUMNS, start=1):
            ws_acc.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ""))
        _style_vdr_row(ws_acc, row_idx, r["_n_rating"], VDR_ACCEPTED_COLUMNS)

    try:
        wb.save(output_xlsx)
    except PermissionError:
        print(f"ERROR: Cannot save — is {output_xlsx} open in Excel?")
        sys.exit(1)

    import json as _json

    def _serialize(val):
        if isinstance(val, (datetime, date)):
            return str(val)
        return val

    vdr_json = {
        "generated": str(today),
        "baseline": baseline,
        "source_poam": str(args.poam),
        "update_frequency_days": VDR_UPDATE_FREQ_DAYS[baseline],
        "active_vulnerabilities": [],
        "accepted_vulnerabilities": [],
    }

    for r in active:
        scheduled = r.get("Scheduled Completion Date") or ""
        vdr_json["active_vulnerabilities"].append({
            "tracking_id":           r.get("POA&M ID", ""),
            "detection_date":        _serialize(r.get("Original Detection Date", "")),
            "detection_source":      r.get("Weakness Detector Source", ""),
            "evaluation_deadline":   r["_eval_deadline"],
            "evaluation_status":     r["_eval_status"],
            "internet_reachability": r["_irv"],
            "likely_exploitability": r["_lev"],
            "cisa_kev":              r["_in_kev"],
            "epss_score":            r["_epss"],
            "adverse_impact":        r["_n_rating"],
            "cve":                   r.get("CVE", ""),
            "weakness_name":         r.get("Weakness Name", ""),
            "asset_identifier":      r.get("Asset Identifier", ""),
            "current_status":        r.get("Comments", "") or "Open",
            "next_milestone":        r.get("Milestone Changes", "") or _serialize(scheduled),
            "scheduled_completion":  _serialize(scheduled),
            "overdue_explanation":   r["_overdue_explanation"],
            "supplementary_risk":    r.get("Overall Remediation Plan", ""),
            "disposition":           "Active",
        })

    for r in accepted:
        is_fp = str(r.get("False Positive") or "").strip().lower() == "yes"
        is_or = str(r.get("Operational Requirement") or "").strip().lower() == "yes"
        age_days = r["_age_days"]
        if is_fp:
            rationale = f"False Positive — {r.get('Deviation Rationale', '')}"
        elif is_or:
            rationale = f"Operational Requirement — {r.get('Deviation Rationale', '')}"
        else:
            rationale = (
                f"Finding exceeded {VDR_ACCEPTANCE_THRESHOLD_DAYS}-day threshold "
                f"({age_days} days). {r.get('Deviation Rationale', '')}"
            )
        vdr_json["accepted_vulnerabilities"].append({
            "tracking_id":           r.get("POA&M ID", ""),
            "detection_date":        _serialize(r.get("Original Detection Date", "")),
            "detection_source":      r.get("Weakness Detector Source", ""),
            "evaluation_deadline":   r["_eval_deadline"],
            "internet_reachability": r["_irv"],
            "likely_exploitability": r["_lev"],
            "cisa_kev":              r["_in_kev"],
            "epss_score":            r["_epss"],
            "adverse_impact":        r["_n_rating"],
            "cve":                   r.get("CVE", ""),
            "weakness_name":         r.get("Weakness Name", ""),
            "asset_identifier":      r.get("Asset Identifier", ""),
            "acceptance_rationale":  rationale.strip(),
            "agency_mitigation":     r.get("Overall Remediation Plan", ""),
        })

    with open(output_json, "w") as f:
        _json.dump(vdr_json, f, indent=2, default=str)

    update_label = {
        "low": "Monthly (every 30 days)",
        "moderate": "Every 14 days",
        "high": "Weekly (every 7 days)",
    }
    next_due = today + timedelta(days=VDR_UPDATE_FREQ_DAYS[baseline])

    w = 62
    print("=" * w)
    print(f"  VDR REPORT — {Path(args.poam).name}")
    print(f"  Baseline: {baseline.upper()} | Generated: {today}")
    print("=" * w)
    print(f"  {'ACTIVE VULNERABILITIES:':<30}{len(active)}")
    print(f"  {'ACCEPTED VULNERABILITIES:':<30}{len(accepted)}")
    print(f"  {'CISA KEV FINDINGS:':<30}{kev_count}  {'<-- Immediate action required' if kev_count else ''}")
    print(f"  {'EVALUATION OVERDUE:':<30}{eval_overdue}")
    print("-" * w)
    print(f"  Required update frequency: {update_label[baseline]}")
    print(f"  Next VDR submission due:   {next_due}")
    print("-" * w)
    if kev_count:
        print("  CISA KEV FINDINGS (Actively Exploited in the Wild)")
        for r in active + accepted:
            if r["_in_kev"]:
                print(f"    {r.get('POA&M ID', '?')}  [{r.get('Original Risk Rating', '?'):<8}]  "
                      f"{(r.get('Weakness Name') or '')[:40]}  CVE: {r.get('CVE', '')}")
        print("-" * w)
    print(f"  Excel: {output_xlsx}")
    print(f"  JSON:  {output_json}  (serve via API for FedRAMP 20x compliance)")
    print("=" * w)


def cmd_vdr_status(args):
    wb_src, ws_open, _ = load_poam(args.poam)
    open_rows = get_all_rows(ws_open)
    today = date.today()
    baseline = args.baseline

    all_cves = [r.get("CVE", "") for r in open_rows if r.get("CVE")]

    print("Checking CISA KEV and EPSS...")
    kev_set = _fetch_cisa_kev()
    epss_scores = _fetch_epss_scores(all_cves)

    active, accepted = _build_vdr_records(open_rows, baseline, kev_set, epss_scores)

    kev_findings = [r for r in active + accepted if r["_in_kev"]]
    eval_overdue = [r for r in active if r["_eval_status"] == "Overdue"]
    rem_overdue = [r for r in active if r["_overdue_explanation"]]
    next_due = today + timedelta(days=VDR_UPDATE_FREQ_DAYS[baseline])

    w = 62
    print("=" * w)
    print(f"  VDR STATUS — {Path(args.poam).name}")
    print(f"  Baseline: {baseline.upper()} | {today}")
    print("=" * w)
    print(f"  {'ACTIVE:':<22}{len(active):<10}{'ACCEPTED:':<18}{len(accepted)}")
    print(f"  {'CISA KEV:':<22}{len(kev_findings):<10}{'EVAL OVERDUE:':<18}{len(eval_overdue)}")
    print(f"  {'REMEDIATION OVERDUE:':<22}{len(rem_overdue):<10}{'NEXT VDR DUE:':<18}{next_due}")
    print("-" * w)

    if kev_findings:
        print("  [!] CISA KEV — Actively Exploited (Immediate Action Required)")
        for r in kev_findings:
            print(f"      {r.get('POA&M ID', '?')}  [{r.get('Original Risk Rating', '?'):<8}]  "
                  f"{(r.get('Weakness Name') or '')[:38]}  {r.get('CVE', '')}")
        print("-" * w)

    if eval_overdue:
        print(f"  [!] EVALUATION DEADLINE OVERDUE ({VDR_EVAL_DEADLINE_DAYS[baseline]}-day window per {baseline} baseline)")
        for r in eval_overdue:
            print(f"      {r.get('POA&M ID', '?')}  [{r.get('Original Risk Rating', '?'):<8}]  "
                  f"Deadline: {r['_eval_deadline']}  {(r.get('Weakness Name') or '')[:30]}")

    print("=" * w)


def main():
    parser = argparse.ArgumentParser(
        prog="grc_tool",
        description="FedRAMP GRC CLI — POA&M management tool",
    )
    sub = parser.add_subparsers(dest="command", required=True)

    p_conv = sub.add_parser("convert", help="Import scanner findings into POA&M")
    p_conv.add_argument("--input", required=True, help="Scanner export file path")
    p_conv.add_argument("--scanner", required=True,
                        choices=["nessus", "tenable", "qualys", "wiz", "generic"])
    p_conv.add_argument("--output", default="poam_output.xlsx", help="Output POA&M file")

    p_enrich = sub.add_parser("enrich", help="Add NIST 800-53 mappings and AI remediation")
    p_enrich.add_argument("--poam", required=True, help="Path to master POA&M file")
    p_enrich.add_argument("--ai", action="store_true",
                          help="Use Ollama/Mistral to generate remediation text")

    p_close = sub.add_parser("close", help="Move a finding to Closed sheet")
    p_close.add_argument("--poam", required=True)
    p_close.add_argument("--id", required=True, dest="poam_id", help="POA&M ID (e.g. POAM-0003)")
    p_close.add_argument("--date", default=date.today().strftime("%Y-%m-%d"), dest="close_date")
    p_close.add_argument("--method", default="Remediated", help="Closure method description")

    p_update = sub.add_parser("update", help="Update fields on an open finding")
    p_update.add_argument("--poam", required=True)
    p_update.add_argument("--id", required=True, dest="poam_id")
    p_update.add_argument("--milestone", default=None, help="Milestone notes")
    p_update.add_argument("--poc", default=None, help="Point of Contact name")
    p_update.add_argument("--vendor-date", default=None, dest="vendor_date",
                          help="Last vendor check-in date (YYYY-MM-DD)")
    p_update.add_argument("--status", default=None, help="Status text (appended to Comments)")

    p_dev = sub.add_parser("deviation", help="Mark finding as False Positive or Operational Requirement")
    p_dev.add_argument("--poam", required=True)
    p_dev.add_argument("--id", required=True, dest="poam_id")
    p_dev.add_argument("--type", required=True, choices=["fp", "or"], dest="dev_type",
                       help="fp = False Positive, or = Operational Requirement")
    p_dev.add_argument("--rationale", required=True, help="Written justification")

    p_dash = sub.add_parser("dashboard", help="Print CLI health dashboard")
    p_dash.add_argument("--poam", required=True)

    p_report = sub.add_parser("report", help="Generate Excel executive summary")
    p_report.add_argument("--poam", required=True)
    p_report.add_argument("--output", default=None, help="Output file (default: grc_report_YYYYMMDD.xlsx)")

    p_conmon = sub.add_parser("conmon", help="Generate monthly ConMon report")
    p_conmon.add_argument("--poam", required=True)
    p_conmon.add_argument("--month", default=None, help="Month in YYYY-MM format (default: current month)")
    p_conmon.add_argument("--output", default=None, help="Output file (default: conmon_YYYY_MM.xlsx)")

    p_export = sub.add_parser("export", help="Export clean FedRAMP POA&M for assessor submission")
    p_export.add_argument("--poam", required=True)
    p_export.add_argument("--output", default=None, help="Output file (default: fedramp_poam_YYYYMMDD.xlsx)")

    p_vdr = sub.add_parser("vdr", help="Generate FedRAMP 20x VDR report (Excel + JSON)")
    p_vdr.add_argument("--poam", required=True)
    p_vdr.add_argument("--baseline", default="moderate", choices=["low", "moderate", "high"],
                       help="FedRAMP baseline (default: moderate)")
    p_vdr.add_argument("--output", default=None, help="Output file (default: vdr_YYYYMMDD.xlsx)")

    p_vdrs = sub.add_parser("vdr-status", help="Quick VDR health check — CISA KEV, eval deadlines, next due date")
    p_vdrs.add_argument("--poam", required=True)
    p_vdrs.add_argument("--baseline", default="moderate", choices=["low", "moderate", "high"])

    args = parser.parse_args()

    dispatch = {
        "convert": cmd_convert,
        "enrich": cmd_enrich,
        "close": cmd_close,
        "update": cmd_update,
        "deviation": cmd_deviation,
        "dashboard": cmd_dashboard,
        "report": cmd_report,
        "conmon": cmd_conmon,
        "export": cmd_export,
        "vdr": cmd_vdr,
        "vdr-status": cmd_vdr_status,
    }
    dispatch[args.command](args)


if __name__ == "__main__":
    main()
